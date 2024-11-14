import openpyxl
import pandas as pd
import tiktoken
from openai import OpenAI
import psycopg2
from pgvector.psycopg2 import register_vector
import numpy as np

# 1. Setup all of the random shit we need. (OpenAI, Database connection, database cursor)
client = OpenAI(
    api_key=''
)

# Set up the connection to the database
conn = psycopg2.connect(
    dbname="mydb",
    user="myuser",
    password="mypassword",
    host="localhost",
    port="5432"
)

# Register pgvector extension
register_vector(conn)

# Create a cursor
cursor = conn.cursor()


def get_embedding(text: str, model="text-embedding-3-small"):
   text = text.replace("\n", " ")
   return client.embeddings.create(input = [text], model=model).data[0].embedding

def get_top3_similar_docs(query_embedding, cur):
    # Get the top 3 most similar documents using the KNN <=> operator
    cur.execute("SELECT text_chunk FROM embeddings ORDER BY embedding <=> %s LIMIT 3", (query_embedding,))
    top3_docs = cur.fetchall()
    return top3_docs

# 1. Open the excel spreadsheet and get the max columns and max rows.
path = "questions.xlsx"

wb_obj = openpyxl.load_workbook(path)

sheet_obj = wb_obj.active

row = sheet_obj.max_row
column = sheet_obj.max_column

# 2. Create the pandas data frame to store data. (question_number, question).
columns = ["Question_Number", "Question"]
data = []
for i in range(2, row + 1):
    current_row = []
    for y in range(1, column+1):
        cell_obj = sheet_obj.cell(row=i, column=y)
        current_row.append(cell_obj.value)
    data.append(current_row)

questions_dataframe = pd.DataFrame(data=data, columns=columns)

# 3. Chunk the transcript by semantic meaning.
max_tokens = 150  # Or 8192 for GPT-4-turbo
transcript_text = ''
with open(file='interview.txt', mode='r', encoding='utf-8') as file:
    transcript_text = file.read()

tokenizer = tiktoken.encoding_for_model('gpt-4o-mini')  # Use "gpt-4" for GPT-4 models if needed
tokens = tokenizer.encode(transcript_text)

chunks = [tokens[i:i + max_tokens] for i in range(0, len(tokens), max_tokens)] # Split tokens into chunks based on the max token limit

text_chunks = [tokenizer.decode(chunk) for chunk in chunks] # Decode tokens back into text

# 4. Turn those chunks into embeddings.

embeddings = []
for chunk in text_chunks:
    embeddings.append([chunk, np.array(get_embedding(text=chunk))])


insert_query = """
    INSERT INTO embeddings (text_chunk, embedding)
    VALUES (%s, %s);
    """
# 5. Add those embeddings to a vector database.
for embedding in embeddings:
    cursor.execute(insert_query, (embedding[0], embedding[1]))

conn.commit()

output_columns = ['Question Number', 'Question', 'Answer To Question']
data = []
# 6. Query each question against the vector database to find the chunks with the closest similarity.
for index, row in questions_dataframe.iterrows():
    question_number = row["Question_Number"]
    question = row["Question"]
    embedding = np.array(get_embedding(text=question))
    top_3 = get_top3_similar_docs(query_embedding=embedding, cur=cursor)
    print(f"Question Number: {question_number}\nQuestion: {question}")

    # 7. Send those chunks of text along with a system prompt to get the answers to the questions if possible.
    completion = client.chat.completions.create(
        model="gpt-4o-mini",
        messages=[
            {"role": "system", "content": "You are a UX researcher who is looking through a transcript of a conversation you had with a user and you're trying to find the answers to the questions. You are required to give exact quotes from the transcript. If the question was never asked, then reply with \'Question Couldn't be Found\'"},
            {
                "role": "user",
                "content": f"Question: {question}: \nResources: {top_3}"
            }
        ]
    )
    question_answer = completion.choices[0].message.content
    output_data = [question_number, question, question_answer]
    data.append(output_data)

# 8. Save the answers to these questions and answers to a spread sheet.
# Create a DataFrame with the data
df = pd.DataFrame(data, columns=output_columns)

cursor.execute("delete from embeddings")
conn.commit()

# Save the DataFrame to a new Excel file
df.to_excel("output.xlsx", index=False)